# Gera a Planilha Finn (.xlsx) pra subir no Google Sheets.
#
# Duas regras que guiaram tudo aqui:
#  1. Tem que funcionar SOZINHA. Ela é vendida à parte, então quem compra e
#     nunca ouviu falar do Finn precisa conseguir usar do zero. A sincronização
#     é camada opcional, nunca pré-requisito.
#  2. Só fórmula que o Google Sheets entende. Nada de recurso exclusivo do
#     Excel — a conversão do Drive descarta em silêncio, e planilha com fórmula
#     quebrada é pior do que planilha sem a fórmula.
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

# ── identidade do Finn (mesmos hex do app) ──
LARANJA   = "F97316"
LARANJA_C = "FFEDD5"
GRAFITE   = "1E293B"
GRAFITE_E = "0F172A"
FUNDO     = "F8F7F4"
BORDA     = "E2E8F0"
CINZA     = "64748B"
VERDE     = "059669"
VERMELHO  = "DC2626"
AMBAR     = "D97706"
BRANCO    = "FFFFFF"

# Fonte única, escrita explicitamente em toda célula (antes só o "Finn." do
# cabeçalho declarava Calibri — o resto ficava no default implícito do
# openpyxl, que POR ACASO também é Calibri, mas sem controle: bastava um dia
# esse default mudar de versão pra versão que a planilha ficaria com duas
# fontes misturadas sem ninguém perceber no código).
FONTE = "Calibri"

MOEDA = 'R$ #,##0.00'
PCT   = '0%'
DATA  = 'dd/mm/yyyy'

LINHAS = 900  # ~2,5 anos de lancamento diario; enxuto pra caber no upload

CAT_DESPESA = ["Alimentação","Transporte","Moradia","Saúde","Educação","Lazer","Vestuário","Outros"]
CAT_RECEITA = ["Salário","Freelance","Aluguel","Venda","Bônus","Investimento","Outros"]
CATEGORIAS  = ["Alimentação","Transporte","Moradia","Saúde","Educação","Lazer","Vestuário",
               "Salário","Freelance","Aluguel","Venda","Bônus","Investimento","Outros"]

wb = Workbook()

def fill(c): return PatternFill("solid", fgColor=c)

def tint(hex_cor, fator=0.85):
    """Clareia uma cor hex misturando com branco (fator 0 = original, 1 = branco).
    Dá cada aba uma versão "lavada" da própria cor de identidade pra fundo de
    cartão e faixa zebrada, em vez de escolher manualmente um pastel pra cada
    uma — bate exatamente com a cor forte usada nos números/rótulos daquela
    aba, garantido por fórmula, não por olho."""
    r, g, b = int(hex_cor[0:2], 16), int(hex_cor[2:4], 16), int(hex_cor[4:6], 16)
    r = int(r + (255 - r) * fator); g = int(g + (255 - g) * fator); b = int(b + (255 - b) * fator)
    return f"{r:02X}{g:02X}{b:02X}"

def borda_fina(cor=BORDA):
    l = Side(style="thin", color=cor)
    return Border(left=l, right=l, top=l, bottom=l)

# Cor de identidade de cada aba — a mesma já usada em sheet_properties.tabColor
# (a cor da abinha lá embaixo do Google Sheets), agora também espalhada pelo
# CONTEÚDO da aba: rótulo de seção, célula editável em destaque, zebrado.
# Antes tudo — em toda aba — usava laranja: a aba de Dívidas tinha destaque
# laranja, Metas tinha destaque laranja, Análises também. Cada aba parecia a
# mesma pele repetida, e a cor da abinha (que já variava) não dizia respeito
# a nada dentro dela. Ligar os dois dá pra cada aba uma identidade visual
# própria dentro da mesma família — não é "tudo laranja", é "cada aba tem a
# sua", com o laranja reservado pra marca (Config, Lançamentos, Início).
ACCENT = {
    "Início": LARANJA, "Lançamentos": LARANJA, "Config": LARANJA,
    "Análises": "0EA5E9", "Limites": AMBAR, "Metas": VERDE,
    "Contas fixas": "8B5CF6", "Dívidas": VERMELHO, "Racha": "0891B2",
}

def titulo_aba(ws, titulo, subtitulo, ncols=8, aba=None):
    """Cabeçalho comum: faixa grafite com a marca (constante — é o selo do
    Finn, não muda de aba pra aba), uma lasca colorida na borda esquerda e um
    fundo lavado da cor da aba por trás do título — é o que dá a cada aba a
    sua própria identidade sem abrir mão da marca."""
    accent = ACCENT.get(aba, LARANJA)
    faixa_accent = Side(style="thick", color=accent)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value="Finn.")
    c.font = Font(name=FONTE, size=22, bold=True, color=BRANCO)
    c.fill = fill(GRAFITE_E)
    c.alignment = Alignment(vertical="center", indent=1)
    c.border = Border(left=faixa_accent)
    ws.row_dimensions[1].height = 42
    for i in range(2, ncols + 1):
        ws.cell(row=1, column=i).fill = fill(GRAFITE_E)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=titulo)
    c.font = Font(name=FONTE, size=13, bold=True, color=GRAFITE)
    c.fill = fill(tint(accent, 0.92))
    c.alignment = Alignment(vertical="center", indent=1)
    c.border = Border(left=faixa_accent)
    ws.row_dimensions[2].height = 24
    for i in range(2, ncols + 1):
        ws.cell(row=2, column=i).fill = fill(tint(accent, 0.92))

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1, value=subtitulo)
    c.font = Font(name=FONTE, size=10, color=CINZA)
    c.fill = fill(tint(accent, 0.92))
    c.alignment = Alignment(vertical="center", indent=1)
    c.border = Border(left=faixa_accent)
    ws.row_dimensions[3].height = 18
    for i in range(2, ncols + 1):
        ws.cell(row=3, column=i).fill = fill(tint(accent, 0.92))

def cabecalho_tabela(ws, linha, valores, larguras=None):
    for i, v in enumerate(valores, start=1):
        c = ws.cell(row=linha, column=i, value=v)
        c.font = Font(name=FONTE, bold=True, size=10, color=BRANCO)
        c.fill = fill(GRAFITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borda_fina()
    ws.row_dimensions[linha].height = 26
    if larguras:
        for i, w in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def zebrado(ws, faixa, cor_base=CINZA):
    """Listra linhas pares com um lavado bem sutil da cor — separa visualmente
    cada linha numa tabela longa sem competir com o conteúdo. stopIfTrue=False
    porque em Lançamentos essa regra convive com outras (Receita/Despesa,
    aviso de linha sem tipo) que também precisam continuar avaliando."""
    ws.conditional_formatting.add(faixa,
        FormulaRule(formula=["ISEVEN(ROW())"], fill=fill(tint(cor_base, 0.94)), stopIfTrue=False))

def rotulo(ws, cell, texto, cor=CINZA, tam=9, bold=True):
    c = ws[cell]; c.value = texto
    c.font = Font(name=FONTE, size=tam, bold=bold, color=cor)

# ══════════════════════════════════════════════════════════════════
#  CONFIG  — precisa existir antes das outras (elas referenciam)
# ══════════════════════════════════════════════════════════════════
cfg = wb.active
cfg.title = "Config"
cfg.sheet_properties.tabColor = CINZA
titulo_aba(cfg, "Configuração", "Mês de referência, categorias e conexão com o app Finn.", 7, aba="Config")

hoje = dt.date.today()
primeiro = dt.date(hoje.year, hoje.month, 1)

rotulo(cfg, "A5", "MÊS DE REFERÊNCIA", LARANJA, 10)
cfg["A6"] = "Mês exibido no Início, Limites e Contas fixas:"
cfg["A6"].font = Font(size=10, color=CINZA)
cfg["D6"] = primeiro
cfg["D6"].number_format = "mmmm/yyyy"
cfg["D6"].font = Font(size=12, bold=True, color=LARANJA)
cfg["D6"].fill = fill(LARANJA_C)
cfg["D6"].border = borda_fina()
cfg["D6"].alignment = Alignment(horizontal="center")
cfg["E6"] = "← troque aqui"
cfg["E6"].font = Font(size=9, italic=True, color=CINZA)

# Chave textual do mês: é ela que as fórmulas SUMIFS comparam. Comparar
# data com data em SUMIFS exige faixa (>= e <=) e quebra fácil; texto
# "aaaa-mm" é exato e não depende de fuso nem de formato regional.
cfg["D7"] = '=TEXT(D6,"yyyy-mm")'
cfg["D7"].font = Font(size=9, color=CINZA)
cfg["D7"].alignment = Alignment(horizontal="center")
cfg["C7"] = "chave (não mexer):"
cfg["C7"].font = Font(size=9, color=CINZA)
cfg["C7"].alignment = Alignment(horizontal="right")

# mês anterior, pra comparação no Início
cfg["D8"] = '=TEXT(EDATE(D6,-1),"yyyy-mm")'
cfg["D8"].font = Font(size=9, color=CINZA)
cfg["D8"].alignment = Alignment(horizontal="center")
cfg["C8"] = "mês anterior:"
cfg["C8"].font = Font(size=9, color=CINZA)
cfg["C8"].alignment = Alignment(horizontal="right")

rotulo(cfg, "A11", "CATEGORIAS", LARANJA, 10)
cfg["A12"] = "Edite à vontade — as listas suspensas da aba Lançamentos seguem daqui."
cfg["A12"].font = Font(size=9, color=CINZA)

cabecalho_tabela(cfg, 13, ["Despesa", "Receita", "Todas (não editar)"],
                 larguras=[22, 22, 24, 14, 16, 14, 14])
for i, v in enumerate(CAT_DESPESA):
    c = cfg.cell(row=14 + i, column=1, value=v); c.border = borda_fina(); c.font = Font(size=10)
for i, v in enumerate(CAT_RECEITA):
    c = cfg.cell(row=14 + i, column=2, value=v); c.border = borda_fina(); c.font = Font(size=10)
for i, v in enumerate(CATEGORIAS):
    c = cfg.cell(row=14 + i, column=3, value=v); c.border = borda_fina(); c.font = Font(size=10, color=CINZA)
zebrado(cfg, f"A14:C{13 + len(CATEGORIAS)}", LARANJA)

rotulo(cfg, "A32", "CONEXÃO COM O APP FINN (opcional)", LARANJA, 10)
cfg["A33"] = ("A planilha funciona 100% sozinha. Preencha abaixo só se você também usa o app Finn "
              "e quer que os lançamentos apareçam nos dois.")
cfg["A33"].font = Font(size=9, color=CINZA)
cfg["A35"] = "Token da planilha"
cfg["A35"].font = Font(size=10, bold=True)
cfg["C35"] = ""
cfg["C35"].fill = fill(LARANJA_C); cfg["C35"].border = borda_fina()
cfg["D35"] = "← cole aqui (Finn → Configurações → Conectar planilha)"
cfg["D35"].font = Font(size=9, italic=True, color=CINZA)
cfg["A36"] = "Última sincronização"
cfg["A36"].font = Font(size=10, bold=True)
cfg["C36"] = ""
cfg["C36"].border = borda_fina()
cfg["C36"].font = Font(size=9, color=CINZA)
cfg["A38"] = "Sem token preenchido, o menu Finn → Sincronizar apenas avisa e não faz nada."
cfg["A38"].font = Font(size=9, italic=True, color=CINZA)

cfg.sheet_view.showGridLines = False

# ══════════════════════════════════════════════════════════════════
#  LANÇAMENTOS — o coração
# ══════════════════════════════════════════════════════════════════
lan = wb.create_sheet("Lançamentos")
lan.sheet_properties.tabColor = LARANJA
titulo_aba(lan, "Lançamentos", "Uma linha por gasto ou entrada. É a única aba onde você digita valores.", 8, aba="Lançamentos")

cabecalho_tabela(lan, 5, ["Data", "Tipo", "Categoria", "Descrição", "Valor",
                          "Mês", "ID Finn", "Origem"],
                 larguras=[13, 12, 18, 38, 15, 11, 24, 12])
lan.freeze_panes = "A6"

for r in range(6, 6 + LINHAS):
    lan.cell(row=r, column=1).number_format = DATA
    lan.cell(row=r, column=5).number_format = MOEDA
    # Chave do mês: vazia enquanto a linha estiver vazia, pra não sujar as
    # somas nem mostrar erro em 2000 linhas em branco.
    lan.cell(row=r, column=6, value=f'=IF($A{r}="","",TEXT($A{r},"yyyy-mm"))')
    lan.cell(row=r, column=6).font = Font(size=9, color=CINZA)
    lan.cell(row=r, column=6).alignment = Alignment(horizontal="center")
    lan.cell(row=r, column=7).font = Font(size=8, color=BORDA)
    lan.cell(row=r, column=8).font = Font(size=9, color=CINZA)
    lan.cell(row=r, column=8).alignment = Alignment(horizontal="center")
    for cidx in range(1, 9):
        lan.cell(row=r, column=cidx).border = borda_fina()

dv_tipo = DataValidation(type="list", formula1='"Despesa,Receita"', allow_blank=True)
dv_tipo.error = "Escolha Despesa ou Receita."
lan.add_data_validation(dv_tipo)
dv_tipo.add(f"B6:B{5 + LINHAS}")

dv_cat = DataValidation(type="list", formula1=f"=Config!$C$14:$C${13 + len(CATEGORIAS)}", allow_blank=True)
dv_cat.error = "Categoria fora da lista da aba Config."
lan.add_data_validation(dv_cat)
dv_cat.add(f"C6:C{5 + LINHAS}")

faixa = f"A6:H{5 + LINHAS}"
lan.conditional_formatting.add(faixa, FormulaRule(formula=[f'$B6="Receita"'],
                                                  fill=fill("ECFDF5"), stopIfTrue=False))
lan.conditional_formatting.add(f"E6:E{5 + LINHAS}",
    FormulaRule(formula=[f'$B6="Despesa"'], font=Font(color=VERMELHO, bold=True)))
lan.conditional_formatting.add(f"E6:E{5 + LINHAS}",
    FormulaRule(formula=[f'$B6="Receita"'], font=Font(color=VERDE, bold=True)))
# Linha preenchida sem tipo é erro de digitação silencioso — some das somas
# e ninguém percebe. Marca em âmbar.
lan.conditional_formatting.add(f"A6:H{5 + LINHAS}",
    FormulaRule(formula=[f'AND($A6<>"",$B6="")'], fill=fill("FEF3C7")))

lan["J5"] = "Confira"
lan["J5"].font = Font(bold=True, size=10, color=GRAFITE)
lan["J6"] = "Linhas preenchidas"
lan["J6"].font = Font(size=9, color=CINZA)
lan["L6"] = f'=COUNTA($A$6:$A${5 + LINHAS})'
lan["J7"] = "Sem tipo (corrigir)"
lan["J7"].font = Font(size=9, color=CINZA)
lan["L7"] = f'=SUMPRODUCT(($A$6:$A${5+LINHAS}<>"")*($B$6:$B${5+LINHAS}=""))'
lan["J8"] = "Sem categoria"
lan["J8"].font = Font(size=9, color=CINZA)
lan["L8"] = f'=SUMPRODUCT(($A$6:$A${5+LINHAS}<>"")*($C$6:$C${5+LINHAS}=""))'
for cel in ["L6", "L7", "L8"]:
    lan[cel].font = Font(size=10, bold=True)
    lan[cel].alignment = Alignment(horizontal="center")
lan.column_dimensions["J"].width = 20
lan.column_dimensions["L"].width = 8
lan.conditional_formatting.add("L7:L8", CellIsRule(operator="greaterThan", formula=["0"],
                                                   font=Font(color=VERMELHO, bold=True)))

L = f"'Lançamentos'!$A$6:$A${5+LINHAS}"
LT = f"'Lançamentos'!$B$6:$B${5+LINHAS}"
LC = f"'Lançamentos'!$C$6:$C${5+LINHAS}"
LV = f"'Lançamentos'!$E$6:$E${5+LINHAS}"
LM = f"'Lançamentos'!$F$6:$F${5+LINHAS}"

# ══════════════════════════════════════════════════════════════════
#  INÍCIO — o dashboard
# ══════════════════════════════════════════════════════════════════
ini = wb.create_sheet("Início", 0)
ini.sheet_properties.tabColor = GRAFITE_E
titulo_aba(ini, "Resumo do mês", "Tudo aqui é calculado sozinho a partir da aba Lançamentos.", 8, aba="Início")
ini.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [22, 16, 4, 22, 16, 4, 22, 16]):
    ini.column_dimensions[col].width = w

ini["A5"] = "=Config!D6"
ini["A5"].number_format = "mmmm/yyyy"
ini["A5"].font = Font(size=14, bold=True, color=GRAFITE)
ini["C5"] = "(troque o mês na aba Config)"
ini["C5"].font = Font(size=9, italic=True, color=CINZA)

def cartao(ws, linha, col, titulo, formula, cor, fmt=MOEDA):
    """Cartão de estatística: fundo lavado da cor do valor + lasca sólida na
    borda esquerda, em vez da caixa branca com contorno fino de antes. A cor
    já dizia o que o número significa (verde=bom, vermelho=ruim); o fundo
    lavado deixa isso visível de relance, sem precisar ler o número."""
    cl = get_column_letter(col); cl2 = get_column_letter(col + 1)
    fino = Side(style="thin", color=BORDA)
    grosso = Side(style="thick", color=cor)
    fundo = fill(tint(cor, 0.9))

    ws.merge_cells(f"{cl}{linha}:{cl2}{linha}")
    c = ws[f"{cl}{linha}"]; c.value = titulo
    c.font = Font(name=FONTE, size=9, bold=True, color=CINZA)
    c.alignment = Alignment(indent=2)
    ws.merge_cells(f"{cl}{linha+1}:{cl2}{linha+1}")
    v = ws[f"{cl}{linha+1}"]; v.value = formula
    v.font = Font(name=FONTE, size=17, bold=True, color=cor)
    v.number_format = fmt
    v.alignment = Alignment(indent=2)

    ws.row_dimensions[linha].height = 19
    ws.row_dimensions[linha + 1].height = 28
    for r in (linha, linha + 1):
        for cc in (col, col + 1):
            cel = ws.cell(row=r, column=cc)
            cel.fill = fundo
            cel.border = Border(
                left=grosso if cc == col else None,
                right=fino if cc == col + 1 else None,
                top=fino if r == linha else None,
                bottom=fino if r == linha + 1 else None,
            )

REC = f'SUMIFS({LV},{LM},Config!$D$7,{LT},"Receita")'
DES = f'SUMIFS({LV},{LM},Config!$D$7,{LT},"Despesa")'
cartao(ini, 7, 1, "RECEITAS DO MÊS", f"={REC}", VERDE)
cartao(ini, 7, 4, "DESPESAS DO MÊS", f"={DES}", VERMELHO)
cartao(ini, 7, 7, "SALDO", f"={REC}-{DES}", GRAFITE)

cartao(ini, 10, 1, "LANÇAMENTOS NO MÊS", f'=COUNTIFS({LM},Config!$D$7)', GRAFITE, fmt="0")
cartao(ini, 10, 4, "GASTO MÉDIO POR DIA",
       f'=IFERROR({DES}/DAY(EOMONTH(Config!$D$6,0)),0)', GRAFITE)
# Taxa de poupança: quanto do que entrou sobrou. É o número que diz se o mês
# prestou — saldo positivo com receita alta pode ainda ser um mês ruim.
cartao(ini, 10, 7, "TAXA DE POUPANÇA",
       f'=IFERROR(({REC}-{DES})/{REC},0)', LARANJA, fmt=PCT)

ini["A14"] = f'=IFERROR(IF({DES}=0,"Sem despesa lançada neste mês.",'\
             f'IF(SUMIFS({LV},{LM},Config!$D$8,{LT},"Despesa")=0,"Sem mês anterior para comparar.",'\
             f'"Você gastou "&TEXT(ABS({DES}/SUMIFS({LV},{LM},Config!$D$8,{LT},"Despesa")-1),"0%")&'\
             f'IF({DES}<SUMIFS({LV},{LM},Config!$D$8,{LT},"Despesa")," a menos"," a mais")&'\
             f'" que no mês anterior.")),"")'
ini.merge_cells("A14:H14")
ini["A14"].font = Font(size=11, bold=True, color=GRAFITE)
ini["A14"].fill = fill(LARANJA_C)
ini["A14"].alignment = Alignment(indent=1, vertical="center")
ini.row_dimensions[14].height = 24

rotulo(ini, "A17", "PARA ONDE FOI O DINHEIRO", LARANJA, 11)
cabecalho_tabela(ini, 18, ["Categoria", "Gasto", "", "% do total", "Barra"])
ini.column_dimensions["C"].width = 4
for i, cat in enumerate(CAT_DESPESA):
    r = 19 + i
    ini.cell(row=r, column=1, value=cat).font = Font(size=10)
    c = ini.cell(row=r, column=2, value=f'=SUMIFS({LV},{LM},Config!$D$7,{LT},"Despesa",{LC},$A{r})')
    c.number_format = MOEDA; c.font = Font(size=10)
    p = ini.cell(row=r, column=4, value=f'=IFERROR($B{r}/{DES},0)')
    p.number_format = PCT; p.font = Font(size=10)
    # Barra em texto: sobrevive a qualquer conversão, ao contrário de gráfico.
    b = ini.cell(row=r, column=5, value=f'=IF($B{r}=0,"",REPT("|",ROUND($D{r}*40,0)))')
    b.font = Font(size=10, color=LARANJA)
    for cc in (1, 2, 4, 5):
        ini.cell(row=r, column=cc).border = borda_fina()
ini.column_dimensions["E"].width = 30

rotulo(ini, "A30", "COMO ESTÃO SEUS LIMITES", LARANJA, 11)
ini["A31"] = '=IF(COUNTA(Limites!$A$6:$A$25)=0,"Nenhum limite definido — vá na aba Limites.","")'
ini["A31"].font = Font(size=9, italic=True, color=CINZA)
cabecalho_tabela(ini, 32, ["Categoria", "Gasto", "", "Limite", "Situação"])
for i in range(8):
    r = 33 + i
    ini.cell(row=r, column=1, value=f'=IF(Limites!$A{6+i}="","",Limites!$A{6+i})').font = Font(size=10)
    c = ini.cell(row=r, column=2, value=f'=IF($A{r}="","",Limites!$C{6+i})')
    c.number_format = MOEDA; c.font = Font(size=10)
    c = ini.cell(row=r, column=4, value=f'=IF($A{r}="","",Limites!$B{6+i})')
    c.number_format = MOEDA; c.font = Font(size=10)
    s = ini.cell(row=r, column=5, value=f'=IF($A{r}="","",IF(Limites!$E{6+i}>1,"Estourou",'
                                        f'IF(Limites!$E{6+i}>0.8,"Atenção","Ok")))')
    s.font = Font(size=10, bold=True)
    for cc in (1, 2, 4, 5):
        ini.cell(row=r, column=cc).border = borda_fina()
ini.conditional_formatting.add("E33:E40",
    FormulaRule(formula=['$E33="Estourou"'], font=Font(color=VERMELHO, bold=True), fill=fill("FEE2E2")))
ini.conditional_formatting.add("E33:E40",
    FormulaRule(formula=['$E33="Atenção"'], font=Font(color=AMBAR, bold=True), fill=fill("FEF3C7")))
ini.conditional_formatting.add("E33:E40",
    FormulaRule(formula=['$E33="Ok"'], font=Font(color=VERDE, bold=True)))

# ══════════════════════════════════════════════════════════════════
#  ANÁLISES
# ══════════════════════════════════════════════════════════════════
ana = wb.create_sheet("Análises")
ana.sheet_properties.tabColor = "0EA5E9"
titulo_aba(ana, "Análises", "Os 12 meses do ano e o comportamento por categoria.", 8, aba="Análises")
ana.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [14, 16, 16, 16, 14, 4, 20, 16]):
    ana.column_dimensions[col].width = w

ana["A5"] = "Ano:"
ana["A5"].font = Font(name=FONTE, size=10, bold=True)
ana["B5"] = hoje.year
ana["B5"].font = Font(name=FONTE, size=12, bold=True, color=ACCENT["Análises"])
ana["B5"].fill = fill(tint(ACCENT["Análises"])); ana["B5"].border = borda_fina()
ana["B5"].alignment = Alignment(horizontal="center")
ana["B5"].number_format = "0"

cabecalho_tabela(ana, 7, ["Mês", "Receitas", "Despesas", "Saldo", "Poupança %"])
for i in range(12):
    r = 8 + i
    ana.cell(row=r, column=1, value=f'=TEXT(DATE($B$5,{i+1},1),"mmm")').font = Font(size=10)
    chave = f'TEXT(DATE($B$5,{i+1},1),"yyyy-mm")'
    c = ana.cell(row=r, column=2, value=f'=SUMIFS({LV},{LM},{chave},{LT},"Receita")')
    c.number_format = MOEDA; c.font = Font(size=10, color=VERDE)
    c = ana.cell(row=r, column=3, value=f'=SUMIFS({LV},{LM},{chave},{LT},"Despesa")')
    c.number_format = MOEDA; c.font = Font(size=10, color=VERMELHO)
    c = ana.cell(row=r, column=4, value=f'=$B{r}-$C{r}')
    c.number_format = MOEDA; c.font = Font(size=10, bold=True)
    c = ana.cell(row=r, column=5, value=f'=IFERROR(($B{r}-$C{r})/$B{r},0)')
    c.number_format = PCT; c.font = Font(size=10)
    for cc in range(1, 6):
        ana.cell(row=r, column=cc).border = borda_fina()
r = 20
ana.cell(row=r, column=1, value="Total").font = Font(name=FONTE, size=10, bold=True)
for cc, letra in [(2, "B"), (3, "C"), (4, "D")]:
    c = ana.cell(row=r, column=cc, value=f'=SUM({letra}8:{letra}19)')
    c.number_format = MOEDA; c.font = Font(name=FONTE, size=10, bold=True)
for cc in range(1, 6):
    ana.cell(row=r, column=cc).fill = fill(tint(ACCENT["Análises"]))
    ana.cell(row=r, column=cc).border = borda_fina()
ana.conditional_formatting.add("D8:D19",
    CellIsRule(operator="lessThan", formula=["0"], font=Font(color=VERMELHO, bold=True)))
zebrado(ana, "A8:E19", ACCENT["Análises"])

graf = BarChart()
graf.type = "col"; graf.style = 10
graf.title = "Receitas x Despesas"
graf.y_axis.title = None; graf.x_axis.title = None
dados = Reference(ana, min_col=2, max_col=3, min_row=7, max_row=19)
cats = Reference(ana, min_col=1, min_row=8, max_row=19)
graf.add_data(dados, titles_from_data=True)
graf.set_categories(cats)
graf.height = 8; graf.width = 18
ana.add_chart(graf, "G7")

rotulo(ana, "A23", "POR CATEGORIA (ANO TODO)", ACCENT["Análises"], 11)
cabecalho_tabela(ana, 24, ["Categoria", "Total gasto", "Média/mês", "% do total"])
for i, cat in enumerate(CAT_DESPESA):
    r = 25 + i
    ana.cell(row=r, column=1, value=cat).font = Font(size=10)
    c = ana.cell(row=r, column=2,
                 value=f'=SUMPRODUCT(({LT}="Despesa")*({LC}=$A{r})*(LEFT({LM},4)=TEXT($B$5,"0000"))*{LV})')
    c.number_format = MOEDA; c.font = Font(size=10)
    c = ana.cell(row=r, column=3, value=f'=IFERROR($B{r}/12,0)')
    c.number_format = MOEDA; c.font = Font(size=10, color=CINZA)
    c = ana.cell(row=r, column=4, value=f'=IFERROR($B{r}/SUM($B$25:$B${24+len(CAT_DESPESA)}),0)')
    c.number_format = PCT; c.font = Font(size=10)
    for cc in range(1, 5):
        ana.cell(row=r, column=cc).border = borda_fina()
zebrado(ana, f"A25:D{24 + len(CAT_DESPESA)}", ACCENT["Análises"])

pizza = PieChart()
pizza.title = "Gastos por categoria"
dados = Reference(ana, min_col=2, min_row=24, max_row=24 + len(CAT_DESPESA))
cats = Reference(ana, min_col=1, min_row=25, max_row=24 + len(CAT_DESPESA))
pizza.add_data(dados, titles_from_data=True)
pizza.set_categories(cats)
pizza.height = 9; pizza.width = 12
ana.add_chart(pizza, "G24")

# ══════════════════════════════════════════════════════════════════
#  LIMITES
# ══════════════════════════════════════════════════════════════════
lim = wb.create_sheet("Limites")
lim.sheet_properties.tabColor = AMBAR
titulo_aba(lim, "Limites por categoria", "Defina o teto mensal. O gasto vem sozinho dos Lançamentos.", 6, aba="Limites")
cabecalho_tabela(lim, 5, ["Categoria", "Limite mensal", "Gasto no mês", "Restante", "% usado", "Situação"],
                 larguras=[20, 16, 16, 16, 12, 16])
lim.freeze_panes = "A6"
for i in range(20):
    r = 6 + i
    lim.cell(row=r, column=2).number_format = MOEDA
    c = lim.cell(row=r, column=3, value=f'=IF($A{r}="","",SUMIFS({LV},{LM},Config!$D$7,{LT},"Despesa",{LC},$A{r}))')
    c.number_format = MOEDA
    c = lim.cell(row=r, column=4, value=f'=IF($A{r}="","",$B{r}-$C{r})')
    c.number_format = MOEDA
    c = lim.cell(row=r, column=5, value=f'=IFERROR(IF($A{r}="","",$C{r}/$B{r}),0)')
    c.number_format = PCT
    c = lim.cell(row=r, column=6, value=f'=IF($A{r}="","",IF($E{r}>1,"Estourou",IF($E{r}>0.8,"Atenção","Ok")))')
    c.font = Font(bold=True)
    for cc in range(1, 7):
        lim.cell(row=r, column=cc).border = borda_fina()
dv_cat2 = DataValidation(type="list", formula1=f"=Config!$A$14:$A${13 + len(CAT_DESPESA)}", allow_blank=True)
lim.add_data_validation(dv_cat2); dv_cat2.add("A6:A25")
lim.conditional_formatting.add("E6:E25", DataBarRule(start_type="num", start_value=0,
                                                     end_type="num", end_value=1, color=LARANJA))
lim.conditional_formatting.add("F6:F25",
    FormulaRule(formula=['$F6="Estourou"'], font=Font(color=VERMELHO, bold=True), fill=fill("FEE2E2")))
lim.conditional_formatting.add("F6:F25",
    FormulaRule(formula=['$F6="Atenção"'], font=Font(color=AMBAR, bold=True), fill=fill("FEF3C7")))
lim.conditional_formatting.add("F6:F25",
    FormulaRule(formula=['$F6="Ok"'], font=Font(color=VERDE, bold=True)))
# Só A-D: E (barra) e F (Estourou/Atenção/Ok) já têm o próprio sinal visual —
# zebrar por cima competiria com esse fundo em vez de só separar linha.
zebrado(lim, "A6:D25", ACCENT["Limites"])

# ══════════════════════════════════════════════════════════════════
#  METAS
# ══════════════════════════════════════════════════════════════════
met = wb.create_sheet("Metas")
met.sheet_properties.tabColor = VERDE
titulo_aba(met, "Metas", "Quanto falta, e quanto guardar por mês pra bater no prazo.", 8, aba="Metas")
cabecalho_tabela(met, 5, ["Meta", "Valor alvo", "Já guardado", "Falta", "%",
                          "Prazo", "Meses restantes", "Guardar por mês"],
                 larguras=[26, 15, 15, 15, 9, 13, 16, 17])
met.freeze_panes = "A6"
for i in range(20):
    r = 6 + i
    met.cell(row=r, column=2).number_format = MOEDA
    met.cell(row=r, column=3).number_format = MOEDA
    c = met.cell(row=r, column=4, value=f'=IF($A{r}="","",MAX(0,$B{r}-$C{r}))'); c.number_format = MOEDA
    c = met.cell(row=r, column=5, value=f'=IFERROR(IF($A{r}="","",MIN(1,$C{r}/$B{r})),0)'); c.number_format = PCT
    met.cell(row=r, column=6).number_format = DATA
    # DATEDIF é aceito pelos dois; MAX(0,...) evita "faltam -3 meses".
    c = met.cell(row=r, column=7,
                 value=f'=IF(OR($A{r}="",$F{r}=""),"",MAX(0,DATEDIF(TODAY(),MAX($F{r},TODAY()),"m")))')
    c.alignment = Alignment(horizontal="center")
    c = met.cell(row=r, column=8,
                 value=f'=IFERROR(IF(OR($A{r}="",$F{r}=""),"",IF($G{r}=0,$D{r},$D{r}/$G{r})),"")')
    c.number_format = MOEDA; c.font = Font(bold=True, color=ACCENT["Metas"])
    for cc in range(1, 9):
        met.cell(row=r, column=cc).border = borda_fina()
met.conditional_formatting.add("E6:E25", DataBarRule(start_type="num", start_value=0,
                                                     end_type="num", end_value=1, color=VERDE))
zebrado(met, "A6:H25", ACCENT["Metas"])

# ══════════════════════════════════════════════════════════════════
#  CONTAS FIXAS
# ══════════════════════════════════════════════════════════════════
fix = wb.create_sheet("Contas fixas")
fix.sheet_properties.tabColor = "8B5CF6"
titulo_aba(fix, "Contas fixas", "O que se repete todo mês. A coluna Situação olha o mês de referência.", 7, aba="Contas fixas")
cabecalho_tabela(fix, 5, ["Descrição", "Tipo", "Categoria", "Valor", "Dia do mês", "Já lançou?", "Situação"],
                 larguras=[30, 12, 18, 15, 12, 13, 18])
fix.freeze_panes = "A6"
for i in range(30):
    r = 6 + i
    fix.cell(row=r, column=4).number_format = MOEDA
    fix.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    # Casa pela descrição dentro do mês de referência: é o que dá pra fazer
    # sem obrigar a pessoa a marcar caixinha na mão todo mês.
    c = fix.cell(row=r, column=6,
                 value=f'=IF($A{r}="","",IF(COUNTIFS({LM},Config!$D$7,'
                       f"'Lançamentos'!$D$6:$D${5+LINHAS},$A{r})>0,\"Sim\",\"Não\"))")
    c.alignment = Alignment(horizontal="center"); c.font = Font(bold=True)
    c = fix.cell(row=r, column=7,
                 value=f'=IF($A{r}="","",IF($F{r}="Sim","Lançada",'
                       f'IF(DAY(TODAY())>$E{r},"Atrasada","A vencer dia "&$E{r})))')
    for cc in range(1, 8):
        fix.cell(row=r, column=cc).border = borda_fina()
dv_tipo2 = DataValidation(type="list", formula1='"Despesa,Receita"', allow_blank=True)
fix.add_data_validation(dv_tipo2); dv_tipo2.add("B6:B35")
dv_cat3 = DataValidation(type="list", formula1=f"=Config!$C$14:$C${13 + len(CATEGORIAS)}", allow_blank=True)
fix.add_data_validation(dv_cat3); dv_cat3.add("C6:C35")
fix.conditional_formatting.add("G6:G35",
    FormulaRule(formula=['$G6="Atrasada"'], font=Font(color=VERMELHO, bold=True), fill=fill("FEE2E2")))
fix.conditional_formatting.add("G6:G35",
    FormulaRule(formula=['$G6="Lançada"'], font=Font(color=VERDE, bold=True)))
# Só A-E: F (Já lançou?) e G (Situação) já têm o próprio fundo condicional.
zebrado(fix, "A6:E35", ACCENT["Contas fixas"])

# ══════════════════════════════════════════════════════════════════
#  DÍVIDAS
# ══════════════════════════════════════════════════════════════════
div = wb.create_sheet("Dívidas")
div.sheet_properties.tabColor = VERMELHO
titulo_aba(div, "Dívidas", "Quanto falta, e em quantos meses acaba no ritmo atual.", 7, aba="Dívidas")
cabecalho_tabela(div, 5, ["Dívida", "Valor total", "Já pago", "Falta", "Juros % a.m.",
                          "Parcela mensal", "Meses p/ quitar"],
                 larguras=[26, 15, 15, 15, 13, 16, 16])
div.freeze_panes = "A6"
for i in range(20):
    r = 6 + i
    for cc in (2, 3, 6):
        div.cell(row=r, column=cc).number_format = MOEDA
    div.cell(row=r, column=5).number_format = '0.00"%"'
    c = div.cell(row=r, column=4, value=f'=IF($A{r}="","",MAX(0,$B{r}-$C{r}))'); c.number_format = MOEDA
    # NPER com juros; sem juros vira divisão simples. Se a parcela não cobre
    # nem o juro do mês, a dívida nunca acaba — e a planilha diz isso em vez
    # de mostrar #NÚM!.
    c = div.cell(row=r, column=7,
                 value=f'=IF(OR($A{r}="",$F{r}=""),"",'
                       f'IF($E{r}=0,IF($F{r}=0,"",ROUNDUP($D{r}/$F{r},0)),'
                       f'IFERROR(ROUNDUP(NPER($E{r}/100,-$F{r},$D{r}),0),"Parcela não cobre os juros")))')
    c.alignment = Alignment(horizontal="center"); c.font = Font(bold=True)
    for cc in range(1, 8):
        div.cell(row=r, column=cc).border = borda_fina()
zebrado(div, "A6:G25", ACCENT["Dívidas"])
div["A28"] = "Total que ainda falta pagar:"
div["A28"].font = Font(size=11, bold=True, color=GRAFITE)
div["D28"] = "=SUM(D6:D25)"
div["D28"].number_format = MOEDA
div["D28"].font = Font(size=12, bold=True, color=VERMELHO)
div["D28"].fill = fill("FEE2E2"); div["D28"].border = borda_fina()

# ══════════════════════════════════════════════════════════════════
#  RACHA
# ══════════════════════════════════════════════════════════════════
rac = wb.create_sheet("Racha")
rac.sheet_properties.tabColor = "0891B2"
titulo_aba(rac, "Racha de contas", "Dividiu com alguém? Registre aqui quanto cada um deve.", 6, aba="Racha")
cabecalho_tabela(rac, 5, ["Data", "Descrição", "Valor total", "Nº de pessoas",
                          "Cada um paga", "Quem já pagou"],
                 larguras=[13, 30, 15, 14, 16, 30])
rac.freeze_panes = "A6"
for i in range(30):
    r = 6 + i
    rac.cell(row=r, column=1).number_format = DATA
    rac.cell(row=r, column=3).number_format = MOEDA
    rac.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    c = rac.cell(row=r, column=5, value=f'=IFERROR(IF($B{r}="","",$C{r}/$D{r}),"")')
    c.number_format = MOEDA; c.font = Font(bold=True)
    for cc in range(1, 7):
        rac.cell(row=r, column=cc).border = borda_fina()
zebrado(rac, "A6:F35", ACCENT["Racha"])

# ── ordem final das abas ──
wb.move_sheet("Config", offset=len(wb.sheetnames))
for nome in ["Início", "Lançamentos", "Análises", "Limites", "Metas", "Contas fixas", "Dívidas", "Racha"]:
    pass

saida = "Planilha-Finn.xlsx"
wb.save(saida)
print("gerado:", saida)
print("abas:", wb.sheetnames)
